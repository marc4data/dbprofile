"""Excel export — one sheet per table, one row per field, check results + notes column."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from dbprofile.report.renderer import (
    CANONICAL_ORDER, CHECK_LABELS, CHECK_SHORT, _score_color,
)

# ── Palette ────────────────────────────────────────────────────────────────
_HEADER_BG   = "1A1A2E"   # dark navy — matches report chrome
_HEADER_FG   = "CDD6F4"   # light lavender
_SUBHDR_BG   = "56779D"   # check header blue
_SUBHDR_FG   = "FFFFFF"
_META_BG     = "F0F2F8"   # light grey for ordinal / type cells

_SEV_FILL: dict[str, str] = {
    "critical": "F38BA8",
    "warn":     "F9E2AF",
    "ok":       "A6E3A1",
    "info":     "89DCEB",
    "na":       "E8EAF0",
}
_SEV_FG: dict[str, str] = {
    "critical": "6B0020",
    "warn":     "7A4000",
    "ok":       "1A5C1A",
    "info":     "005F78",
    "na":       "8A8FB0",
}

# Score colours (hex, no #)
_SCORE_COLOR: dict[str, str] = {
    "green":  "A6E3A1",
    "yellow": "F9E2AF",
    "orange": "FAB387",
    "red":    "F38BA8",
}

def _score_fill(score: int) -> str:
    if score >= 90: return _SCORE_COLOR["green"]
    if score >= 75: return _SCORE_COLOR["yellow"]
    if score >= 60: return _SCORE_COLOR["orange"]
    return _SCORE_COLOR["red"]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "1A1A2E", size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _thin_border() -> Border:
    s = Side(style="thin", color="D0D4E8")
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


# ── Sheet builders ──────────────────────────────────────────────────────────

def _write_cover(wb: Workbook, context: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    def _row(label: str, value: str, row: int) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=True, color="56779D")
        lc.alignment = _left()
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = _font()
        vc.alignment = _left()

    ws.cell(row=1, column=1, value="dbprofile — Profiling Workbook").font = Font(
        bold=True, size=14, color=_HEADER_BG, name="Calibri"
    )
    ws.row_dimensions[1].height = 24

    tables = context.get("tables", [])
    tables_ctx = context.get("tables_ctx", {})

    _row("Generated",       context.get("run_at", ""),              3)
    _row("Connector",       context.get("dialect_display", ""),      4)
    _row("Account",         context.get("account_display", ""),      5)
    _row("Database",        context.get("database_display", ""),     6)
    _row("Schema",          context.get("schema_display", ""),       7)
    _row("Sample Rate",     context.get("sample_rate_pct", ""),      8)
    _row("Tables Profiled", str(len(tables)),                        9)
    _row("Overall Score",   str(context.get("overall_quality_score", "")), 10)

    # Per-table score summary
    ws.cell(row=12, column=1, value="Table").font = _font(bold=True, color=_SUBHDR_FG)
    ws.cell(row=12, column=1).fill = _fill(_SUBHDR_BG)
    ws.cell(row=12, column=2, value="Quality Score").font = _font(bold=True, color=_SUBHDR_FG)
    ws.cell(row=12, column=2).fill = _fill(_SUBHDR_BG)
    for col_i in range(3, 3 + len(CANONICAL_ORDER)):
        cn = CANONICAL_ORDER[col_i - 3]
        c = ws.cell(row=12, column=col_i, value=CHECK_SHORT[cn])
        c.font = _font(bold=True, color=_SUBHDR_FG)
        c.fill = _fill(_SUBHDR_BG)
        c.alignment = _center()
        c.comment = None
        ws.column_dimensions[get_column_letter(col_i)].width = 8
    ws.column_dimensions[get_column_letter(2 + len(CANONICAL_ORDER) + 1)].width = 8

    for i, table in enumerate(tables):
        tc = tables_ctx.get(table, {})
        r = 13 + i
        name_cell = ws.cell(row=r, column=1, value=table)
        name_cell.font = _font()
        name_cell.alignment = _left()
        score = tc.get("quality_score", "")
        sc = ws.cell(row=r, column=2, value=score)
        sc.font = _font(bold=True)
        sc.alignment = _center()
        if isinstance(score, int):
            sc.fill = _fill(_score_fill(score))
        for col_i, cn in enumerate(CANONICAL_ORDER, start=3):
            sev = tc.get("check_worst", {}).get(cn, "na")
            c = ws.cell(row=r, column=col_i, value=sev)
            c.font = _font(color=_SEV_FG.get(sev, "1A1A2E"), size=9)
            c.fill = _fill(_SEV_FILL.get(sev, "E8EAF0"))
            c.alignment = _center()


def _write_table_sheet(wb: Workbook, table: str, tc: dict[str, Any],
                        sample_rate_pct: str) -> None:
    """One sheet per table: header rows + one data row per column."""
    safe_name = table[:31]  # Excel sheet name limit
    ws = wb.create_sheet(title=safe_name)

    scorecard: list[dict] = tc.get("scorecard", [])
    check_worst: dict[str, str] = tc.get("check_worst", {})
    quality_score: int = tc.get("quality_score", 0)
    row_count: int = tc.get("row_count", 0)
    rows_sampled: int = tc.get("rows_sampled", 0)
    col_count: int = tc.get("col_count", 0)

    # ── Title block ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1, value=f"Table: {table}")
    title.font = Font(bold=True, size=13, color=_HEADER_FG, name="Calibri")
    title.fill = _fill(_HEADER_BG)
    title.alignment = _left()
    ws.row_dimensions[1].height = 20

    # Meta row
    meta_labels = [
        ("Rows", f"{row_count:,}"),
        ("Sampled", f"{rows_sampled:,}  ({sample_rate_pct})"),
        ("Columns", str(col_count)),
        ("Quality Score", str(quality_score)),
    ]
    col_cursor = 1
    for label, val in meta_labels:
        lc = ws.cell(row=2, column=col_cursor, value=label)
        lc.font = _font(bold=True, color="56779D", size=9)
        lc.fill = _fill("EEF1FA")
        lc.alignment = _center()
        vc = ws.cell(row=2, column=col_cursor + 1, value=val)
        vc.font = _font(size=9)
        vc.fill = _fill("EEF1FA")
        vc.alignment = _center()
        if label == "Quality Score" and isinstance(quality_score, int):
            vc.fill = _fill(_score_fill(quality_score))
            vc.font = _font(bold=True, size=9)
        col_cursor += 2
    ws.row_dimensions[2].height = 16

    # Check-level worst-severity row
    # Columns: A=#, B=Field, C=Type, D=Grp, E=Sub, F=Seq, G=EDA, then checks
    eda_col_start = 4    # D
    check_col_start = 8  # H (after 4 EDA columns)
    ws.cell(row=3, column=1, value="Check Results").font = _font(bold=True, color=_SUBHDR_FG, size=9)
    ws.cell(row=3, column=1).fill = _fill(_SUBHDR_BG)
    ws.cell(row=3, column=2, value="(worst per check across all columns)").font = _font(color=_SUBHDR_FG, size=9)
    ws.cell(row=3, column=2).fill = _fill(_SUBHDR_BG)
    for col_i in range(3, check_col_start):
        ws.cell(row=3, column=col_i, value="").fill = _fill(_SUBHDR_BG)
    for i, cn in enumerate(CANONICAL_ORDER):
        sev = check_worst.get(cn, "na")
        c = ws.cell(row=3, column=check_col_start + i, value=sev)
        c.font = _font(bold=True, color=_SEV_FG.get(sev, "1A1A2E"), size=9)
        c.fill = _fill(_SEV_FILL.get(sev, "E8EAF0"))
        c.alignment = _center()

    # Notes column index
    notes_col = check_col_start + len(CANONICAL_ORDER)

    # ── Column headers ───────────────────────────────────────────────────────
    HDR_ROW = 4
    headers = (
        ["#", "Field Name", "Data Type",
         "Grp", "Sub", "Seq", "EDA"] +
        [f"{CHECK_SHORT[cn]}\n{CHECK_LABELS[cn]}" for cn in CANONICAL_ORDER] +
        ["Notes"]
    )

    for col_i, hdr in enumerate(headers, start=1):
        c = ws.cell(row=HDR_ROW, column=col_i, value=hdr)
        c.font = _font(bold=True, color=_HEADER_FG, size=9)
        c.fill = _fill(_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[HDR_ROW].height = 30

    # ── Hidden separator row — enables Excel sort/filter on the data below ──
    SEP_ROW = 5
    ws.row_dimensions[SEP_ROW].hidden = True
    ws.row_dimensions[SEP_ROW].height = 2

    # ── Data rows ────────────────────────────────────────────────────────────
    DATA_START = 6
    for row_i, row in enumerate(scorecard, start=DATA_START):
        ws.row_dimensions[row_i].height = 16

        # Ordinal
        c = ws.cell(row=row_i, column=1, value=row.get("ordinal", ""))
        c.font = _font(color="6C7086", size=9)
        c.fill = _fill(_META_BG)
        c.alignment = _center()
        c.border = _thin_border()

        # Field name
        c = ws.cell(row=row_i, column=2, value=row.get("column", ""))
        c.font = _font(bold=True, size=10)
        c.alignment = _left()
        c.border = _thin_border()

        # Data type
        c = ws.cell(row=row_i, column=3, value=row.get("data_type", ""))
        c.font = _font(color="4C4F69", size=9)
        c.fill = _fill(_META_BG)
        c.alignment = _center()
        c.border = _thin_border()

        # EDA columns
        for eda_i, eda_key in enumerate(["eda_grp", "eda_sub", "eda_seq", "eda_sort"]):
            c = ws.cell(row=row_i, column=eda_col_start + eda_i, value=row.get(eda_key, ""))
            c.font = _font(color="4C4F69", size=9)
            c.fill = _fill(_META_BG)
            c.alignment = _center()
            c.border = _thin_border()

        # Check severity cells
        for i, cn in enumerate(CANONICAL_ORDER):
            sev = row.get(cn, "na")
            c = ws.cell(row=row_i, column=check_col_start + i, value=sev)
            c.font = _font(color=_SEV_FG.get(sev, "1A1A2E"), size=9)
            c.fill = _fill(_SEV_FILL.get(sev, "E8EAF0"))
            c.alignment = _center()
            c.border = _thin_border()

        # Notes column — blank, ready for manual input
        c = ws.cell(row=row_i, column=notes_col, value="")
        c.border = _thin_border()
        c.fill = _fill("FFFDE7")   # very light yellow — invites editing

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5    # #
    ws.column_dimensions["B"].width = 28   # field name
    ws.column_dimensions["C"].width = 18   # data type
    ws.column_dimensions["D"].width = 5    # Grp
    ws.column_dimensions["E"].width = 5    # Sub
    ws.column_dimensions["F"].width = 5    # Seq
    ws.column_dimensions["G"].width = 7    # EDA sort key
    for i in range(len(CANONICAL_ORDER)):
        ws.column_dimensions[get_column_letter(check_col_start + i)].width = 10
    ws.column_dimensions[get_column_letter(notes_col)].width = 40

    # Freeze panes — keep header visible while scrolling (freeze below data header row)
    ws.freeze_panes = ws.cell(row=DATA_START, column=check_col_start)


# ── Public entry point ──────────────────────────────────────────────────────

def write_excel(
    output_path: str | Path,
    context: dict[str, Any],
) -> Path:
    """Write a profiling workbook and return the path written."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_cover(wb, context)

    tables = context.get("tables", [])
    tables_ctx = context.get("tables_ctx", {})
    sample_rate_pct = context.get("sample_rate_pct", "")

    for table in tables:
        tc = tables_ctx.get(table, {})
        _write_table_sheet(wb, table, tc, sample_rate_pct)

    wb.save(output_path)
    return output_path
